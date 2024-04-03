# -*- coding: utf-8 -*-
"""
Created on Sun Mar 24 12:11:26 2024

@author: Lakshitha S
"""

#import and input
import openpyxl,docx,random
from tabulate import tabulate

inputTeachers = "C:\\Users\\Lakshitha S\\Documents\\Clg Wrk\\Timetable\\Teachers.xlsx"
inputCourses = "C:\\Users\\Lakshitha S\\Documents\\Clg Wrk\\Timetable\\Course Sub.xlsx"
xloutput = "C:\\Users\\Lakshitha S\\Documents\\Clg Wrk\\Timetable\\Timetable.docx"

# Creating table layout
table = [ [ ['Mon'],[],[],['b'],[],[],[] ],
          [ ['Tue'],[],[],['r'],[],[],[] ],
          [ ['Wed'],[],[],['e'],[],[],[] ],
          [ ['Thu'],[],[],['a'],[],[],[] ],
          [ ['Fri'],[],[],['k'],[],[],[] ] ]
head = [ ["1"],["2"],["Lunch"],["3"],["4"],["5"] ]
print(tabulate(table, headers=head, tablefmt="grid"))


# Manipulation xl worksheet
wbT = openpyxl.load_workbook(inputTeachers)
wsT = wbT.active
dnry={}
# sorting teachers n the subs they teach, and put them into sorted lists
for r in range(2,wsT.max_row+1):    #Start from 2 cuz 1st row consists of teachers names
    for c in range(2,wsT.max_column+1):     #Start from 2 cuz 1st col consists of headings
        subject = wsT.cell(row=r ,column=c).value
        teacher = wsT.cell(row=r ,column=1).value
        #(dnry[subject].append(teacher)) if (subject in dnry) else (dnry[subject]=[teacher])
        if subject in dnry:
            dnry[subject].append(teacher)
        else:
            dnry[subject]=[teacher]
wbC = openpyxl.load_workbook(inputCourses)
wsC = wbC.active
subsPerCourse = []
for ro in range(2,wsC.max_row+1):    #Start from 2 cuz 1st row consists of teachers names
    course = wsT.cell(row=r ,column=1).value
    print(course)
    for co in range(2,wsC.max_column+1):     #Start from 2 cuz 1st col consists of headings 
        subC = wsT.cell(row=r ,column=c).value   
        subsPerCourse.append(subC)
#Tick of a teacher once assigned in the copy (of the original dictionary of teachers)
print(subsPerCourse)
RemProf = dnry.copy()
for Course_Sub in subsPerCourse:
    assigned_teacher = random.choice(RemProf[Course_Sub])
    print('Course: ', Course_Sub,', Professor:',assigned_teacher)
    
    
        

# print(table)

doc = docx.Document()
table = doc.add_table(rows=1 , columns=2)
